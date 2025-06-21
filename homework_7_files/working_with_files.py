import time

from pypdf import PdfReader
from zipfile import ZipFile
from openpyxl import load_workbook
import csv


def test_working_with_pdf_file(create_archive):
    with ZipFile(create_archive) as zip_file:
        pdf_file = zip_file.open("КП_выпускные_2025.pdf")
        reader = PdfReader(pdf_file)
        page_text = reader.pages[0].extract_text()
        expected_text = "ВЫПУСКНЫЕ\n2025\nДЛЯ ШКОЛ, ДЕТСКИХ САДОВ,\nЧАСТНЫХ ВЕЧЕРИНОК\nАГЕНТСТВО ПРАЗДНИКОВ"

        assert expected_text in page_text


def test_working_with_xlsx_file(create_archive):
    with ZipFile(create_archive) as zip_file:
        xlsx_file = zip_file.open("file_example_XLSX_50.xlsx")
        sheet = load_workbook(xlsx_file).active
        page_text = sheet.cell(row=2, column=3).value
        expected_text = "Abril"

        assert expected_text in page_text


def test_working_with_csv_file(create_archive):
    with ZipFile(create_archive) as zip_file:
        csv_file = zip_file.open("users.csv")
        content = csv_file.read().decode('utf-8-sig')
        csvreader = list(csv.reader(content.splitlines()))
        page_text = csvreader[2]
        expected_text = "autotest1@autotest.clients"

        assert expected_text in page_text
